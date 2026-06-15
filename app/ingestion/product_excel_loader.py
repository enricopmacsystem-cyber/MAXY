from __future__ import annotations

import re
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.core.exceptions import ProductImportError
from app.core.logging import get_logger

logger = get_logger(__name__)

COLUMN_ALIASES: dict[str, str] = {
    "codice_interno": "internal_code",
    "codice": "internal_code",
    "codice articolo": "internal_code",
    "internal_code": "internal_code",
    "produttore": "manufacturer",
    "marca": "manufacturer",
    "manufacturer": "manufacturer",
    "descrizione": "description",
    "description": "description",
    "categoria": "category",
    "category": "category",
    "disponibilita": "availability",
    "disponibilità": "availability",
    "availability": "availability",
    "stock": "availability",
    "giacenza": "availability",
    "prezzo": "price",
    "price": "price",
    "costo": "cost_price",
    "costo_acquisto": "cost_price",
    "cost_price": "cost_price",
    "link_manuale": "manual_url",
    "manuale": "manual_url",
    "manual_url": "manual_url",
    "link manuale": "manual_url",
    "link_scheda_tecnica": "datasheet_url",
    "scheda_tecnica": "datasheet_url",
    "datasheet_url": "datasheet_url",
    "link scheda tecnica": "datasheet_url",
}


def _normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = text.replace("_", " ")
    text = re.sub(r"\s+", " ", text)
    return text


def _map_headers(headers: list[Any]) -> dict[int, str]:
    mapping: dict[int, str] = {}
    for index, header in enumerate(headers):
        normalized = _normalize_header(header)
        normalized_key = normalized.replace(" ", "_")
        field = COLUMN_ALIASES.get(normalized) or COLUMN_ALIASES.get(normalized_key)
        if field:
            mapping[index] = field
    return mapping


def _parse_availability(value: Any) -> int:
    if value is None or str(value).strip() == "":
        return 0
    try:
        number = int(float(str(value).replace(",", ".")))
    except (ValueError, TypeError) as exc:
        raise ProductImportError(f"Disponibilità non valida: {value}") from exc
    if number < 0:
        raise ProductImportError(f"Disponibilità negativa non consentita: {number}")
    return number


def _parse_price(value: Any) -> Decimal:
    if value is None or str(value).strip() == "":
        raise ProductImportError("Prezzo mancante")
    raw = str(value).strip().replace("€", "").replace(" ", "")
    if "," in raw and "." in raw:
        raw = raw.replace(".", "").replace(",", ".")
    else:
        raw = raw.replace(",", ".")
    try:
        price = Decimal(raw)
    except (InvalidOperation, ValueError) as exc:
        raise ProductImportError(f"Prezzo non valido: {value}") from exc
    if price < 0:
        raise ProductImportError(f"Prezzo negativo non consentito: {price}")
    return price.quantize(Decimal("0.01"))


def _parse_optional_url(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _parse_optional_price(value: Any) -> Decimal | None:
    if value is None or str(value).strip() == "":
        return None
    return _parse_price(value)


def _parse_required_text(value: Any, field_name: str) -> str:
    text = str(value or "").strip()
    if not text:
        raise ProductImportError(f"Campo obbligatorio mancante: {field_name}")
    return text


def parse_products_excel(source: Path | BytesIO) -> tuple[list[dict[str, Any]], list[str]]:
    """
    Legge un file Excel e restituisce righe prodotto normalizzate.

    Returns:
        Tuple (rows, errors)
    """
    logger.info("Lettura Excel prodotti: %s", getattr(source, "name", "upload"))

    try:
        workbook = load_workbook(filename=source, read_only=True, data_only=True)
    except Exception as exc:
        raise ProductImportError(f"Impossibile aprire il file Excel: {exc}") from exc

    sheet = workbook.active
    if sheet is None:
        raise ProductImportError("Il file Excel non contiene fogli")

    rows_iter = sheet.iter_rows(values_only=True)
    try:
        headers = next(rows_iter)
    except StopIteration as exc:
        raise ProductImportError("Il file Excel è vuoto") from exc

    header_map = _map_headers(list(headers))
    required_fields = {
        "internal_code",
        "manufacturer",
        "description",
        "category",
        "price",
    }
    missing = required_fields - set(header_map.values())
    if missing:
        raise ProductImportError(
            "Colonne obbligatorie mancanti nel file Excel: "
            + ", ".join(sorted(missing))
        )

    parsed_rows: list[dict[str, Any]] = []
    errors: list[str] = []

    for row_number, row in enumerate(rows_iter, start=2):
        if not row or all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        payload: dict[str, Any] = {}
        try:
            for col_index, field_name in header_map.items():
                value = row[col_index] if col_index < len(row) else None
                payload[field_name] = value

            product = {
                "internal_code": _parse_required_text(
                    payload.get("internal_code"), "codice_interno"
                ),
                "manufacturer": _parse_required_text(
                    payload.get("manufacturer"), "produttore"
                ),
                "description": _parse_required_text(
                    payload.get("description"), "descrizione"
                ),
                "category": _parse_required_text(payload.get("category"), "categoria"),
                "availability": _parse_availability(payload.get("availability")),
                "price": _parse_price(payload.get("price")),
                "cost_price": _parse_optional_price(payload.get("cost_price")),
                "manual_url": _parse_optional_url(payload.get("manual_url")),
                "datasheet_url": _parse_optional_url(payload.get("datasheet_url")),
            }
            parsed_rows.append(product)
        except ProductImportError as exc:
            message = f"Riga {row_number}: {exc}"
            logger.warning(message)
            errors.append(message)

    workbook.close()
    logger.info(
        "Excel analizzato: %d righe valide, %d errori",
        len(parsed_rows),
        len(errors),
    )
    return parsed_rows, errors
