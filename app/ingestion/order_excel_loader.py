from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.core.exceptions import OrderImportError
from app.core.logging import get_logger

logger = get_logger(__name__)

COLUMN_ALIASES: dict[str, str] = {
    "numero_ordine": "order_number",
    "order_number": "order_number",
    "ordine": "order_number",
    "data_ordine": "order_date",
    "order_date": "order_date",
    "data": "order_date",
    "codice_cliente": "customer_code",
    "customer_code": "customer_code",
    "cliente": "customer_code",
    "codice_interno": "internal_code",
    "codice articolo": "internal_code",
    "codice": "internal_code",
    "internal_code": "internal_code",
    "quantita": "quantity",
    "quantità": "quantity",
    "quantity": "quantity",
    "qta": "quantity",
    "prezzo_unitario": "unit_price",
    "prezzo": "unit_price",
    "unit_price": "unit_price",
}


@dataclass(frozen=True)
class ParsedOrderLine:
    order_number: str
    order_date: date
    customer_code: str | None
    internal_code: str
    quantity: Decimal
    unit_price: Decimal | None


def _normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = text.replace("_", " ")
    text = re.sub(r"\s+", " ", text)
    return text


def _map_headers(headers: list[Any]) -> dict[int, str]:
    mapping: dict[int, str] = {}
    for index, header in enumerate(headers):
        normalized = _normalize_header(header)
        normalized_key = normalized.replace(" ", "_")
        field = COLUMN_ALIASES.get(normalized) or COLUMN_ALIASES.get(normalized_key)
        if field:
            mapping[index] = field
    return mapping


def _parse_date(value: Any) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = str(value or "").strip()
    if not text:
        raise OrderImportError("Data ordine mancante")

    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue

    raise OrderImportError(f"Data ordine non valida: {value}")


def _parse_quantity(value: Any) -> Decimal:
    if value is None or str(value).strip() == "":
        raise OrderImportError("Quantità mancante")
    raw = str(value).strip().replace(",", ".")
    try:
        quantity = Decimal(raw)
    except InvalidOperation as exc:
        raise OrderImportError(f"Quantità non valida: {value}") from exc
    if quantity <= 0:
        raise OrderImportError(f"Quantità deve essere > 0: {quantity}")
    return quantity.quantize(Decimal("0.01"))


def _parse_optional_price(value: Any) -> Decimal | None:
    if value is None or str(value).strip() == "":
        return None
    raw = str(value).strip().replace("€", "").replace(" ", "").replace(",", ".")
    try:
        price = Decimal(raw)
    except InvalidOperation as exc:
        raise OrderImportError(f"Prezzo non valido: {value}") from exc
    return price.quantize(Decimal("0.01"))


def _parse_required_text(value: Any, field_name: str) -> str:
    text = str(value or "").strip()
    if not text:
        raise OrderImportError(f"Campo obbligatorio mancante: {field_name}")
    return text


def parse_orders_excel(source: Path | BytesIO) -> tuple[list[ParsedOrderLine], list[str]]:
    logger.info("Lettura Excel ordini: %s", getattr(source, "name", "upload"))

    try:
        workbook = load_workbook(filename=source, read_only=True, data_only=True)
    except Exception as exc:
        raise OrderImportError(f"Impossibile aprire il file Excel: {exc}") from exc

    sheet = workbook.active
    if sheet is None:
        raise OrderImportError("Il file Excel non contiene fogli")

    rows_iter = sheet.iter_rows(values_only=True)
    try:
        headers = next(rows_iter)
    except StopIteration as exc:
        raise OrderImportError("Il file Excel è vuoto") from exc

    header_map = _map_headers(list(headers))
    required = {"order_number", "order_date", "internal_code", "quantity"}
    missing = required - set(header_map.values())
    if missing:
        raise OrderImportError(
            "Colonne obbligatorie mancanti: " + ", ".join(sorted(missing))
        )

    parsed: list[ParsedOrderLine] = []
    errors: list[str] = []

    for row_number, row in enumerate(rows_iter, start=2):
        if not row or all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        payload: dict[str, Any] = {}
        try:
            for col_index, field_name in header_map.items():
                payload[field_name] = row[col_index] if col_index < len(row) else None

            parsed.append(
                ParsedOrderLine(
                    order_number=_parse_required_text(
                        payload.get("order_number"), "numero_ordine"
                    ),
                    order_date=_parse_date(payload.get("order_date")),
                    customer_code=(
                        str(payload.get("customer_code")).strip()
                        if payload.get("customer_code") not in (None, "")
                        else None
                    ),
                    internal_code=_parse_required_text(
                        payload.get("internal_code"), "codice_interno"
                    ),
                    quantity=_parse_quantity(payload.get("quantity")),
                    unit_price=_parse_optional_price(payload.get("unit_price")),
                )
            )
        except OrderImportError as exc:
            message = f"Riga {row_number}: {exc}"
            logger.warning(message)
            errors.append(message)

    workbook.close()
    logger.info("Excel ordini: %d righe valide, %d errori", len(parsed), len(errors))
    return parsed, errors
